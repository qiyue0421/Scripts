# 本脚本用于处理EXCEL文档，作为一个简单的示例
from openpyxl import load_workbook, Workbook
import json

# 加载已存在的文件
wb = load_workbook(filename=r'../test/MQ.xlsx')
# 选择第一个sheet
sh = wb[wb.sheetnames[0]]

# 开启一个工作簿
wb2 = Workbook()
# 激活工作簿
ws = wb2.active
# 写入列名
ws.append(['path', '@timestamp', 'host', 'error_style', 'message'])
# 从第二行开始遍历数据
for cases in list(sh.rows)[1:]:
    # 将字符数据转换成字典类型
    case_content = json.loads(cases[1].value)
    error_style = case_content['message'].split('-')[2].split()[3].split('_')[0].strip()
    # 调用append方法将数据写入空的EXCEL文件
    ws.append([case_content['path'], case_content['@timestamp'], case_content['host'], error_style, case_content['message']])
# 保存工作表
wb2.save('result.xlsx')
# 关闭
wb2.close()
wb.close()
